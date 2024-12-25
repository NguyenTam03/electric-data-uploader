import openpyxl
from tkinter.ttk import *
from tkinter import *
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from tkinter import filedialog
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import chromedriver_autoinstaller 
chrome_profile_directory = r"C:\Users\Administrator\AppData\Local\Google\Chrome\User Data\Default"
file_path = ''
global str_sheet  
global checkbox_vars
global checkbox_values
def login(tk,mk,link_url,xpath_file):
    global str_sheet
    global checkbox_vars
    
    link_cbm =""
    if link_url == 'https://pmis.npc.com.vn/PMIS_Web/tmsLogin.jsf':
        link_cbm = 'https://pmis.npc.com.vn/PMIS_Web/eam/cbm/cbmWork.jsf'
    else:
        link_cbm = 'https://10.21.50.29/PMIS_Web/eam/cbm/cbmWork.jsf'
        
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument('--ignore-certificate-errors')
    options.add_argument(f"--user-data-dir={chrome_profile_directory}")
    driver = webdriver.Chrome(options=options)
    driver.get(link_url)
    time.sleep(2)
    user = driver.find_element(By.ID,"formMain:txtUserName")
    user.send_keys(tk)
    time.sleep(0.5)
    password = driver.find_element(By.ID,"formMain:txtPassWord")
    password.send_keys(mk)
    password.send_keys(Keys.ENTER)
    time.sleep(2)
    try:
        driver.find_element(By.XPATH,'/html/body/div/div[2]/button[3]').click()
    except:
        pass
    time.sleep(2)
    
    driver.get(link_cbm)
    time.sleep(5)
    for i, var in enumerate(checkbox_vars):
        # print(str_sheet[i][0],var.get())
        if var.get():
            # sheet | array_column_part | toltal_part
            fill_a_form(xpath_file, driver, link_cbm, str_sheet[i][0],str_sheet[i][1],len(str_sheet[i][1])-1)
    messagebox.showinfo("Thông Báo", "Hoàn Thành!!!")

def check_None(arr):
    #Check mang 1 chieu
    if any(element == 'None' for element in arr):
        return True
    else:
        return False
    
def fill_a_form(xpath_file, driver, link_cbm, name_sheet, array_column_part, toltal_part):
    print('Đang thực hiện sheet',name_sheet)
    ex = openpyxl.load_workbook(xpath_file)
    sheet = ex[name_sheet]
    info = list()
    temp = ''
    a = list()
    b = list()
    for row in range(3, sheet.max_row+1):
        if sheet[row][0].value is None:
            break
        for col in range(0, sheet.max_column):
            temp += str(sheet[row][col].value) + '|'
            if col in array_column_part:
                a = temp.strip().split('|')
                temp =''
                b.append(a)

        temp=''
        info.append(b)
        b = list()
        
    for row in range(0,len(info)):
        print('Dòng',row+1)
        driver.get(link_cbm)
        time.sleep(1)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/div[4]/div[1]/div/table/tbody/tr/td[3]/input'))).send_keys(info[row][0])
        time.sleep(0.5)
        
        driver.find_element(By.XPATH,'/html/body/form/div[4]/div[1]/div/table/tbody/tr/td[4]/button/span[1]').click()
        time.sleep(1)
        while True:
            time.sleep(1)
            try:
                driver.find_element(By.XPATH,'/html/body/div[3]/div').click()
            except:
                break
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/div[4]/div[2]/div/div/div[2]/table/tbody/tr/td[2]'))).click()
        time.sleep(2)
        # Loop fill part
        # sẽ bị giảm khi để trống (không có dòng có ở web)
        pos_current_part = 0
        # var ko bị giảm
        var_current_part = 0
        while var_current_part < toltal_part:
            var_current_part += 1
            pos_current_part += 1
            print(f"Phần {var_current_part}")
            #----------------------------------------------------------------------------
            print(info[row][var_current_part][0].strip().lower())
            print(var_current_part,toltal_part+1)
            if info[row][var_current_part][0].strip().lower() == 'không nhập':
                sheet[row+3][array_column_part[-1]+var_current_part+1].value = "Không thực hiện"
                ex.save(xpath_file)
                continue
            #----------------------------------------------------------------------------
            elif check_None(info[row][var_current_part]) == True:
                pos_current_part -= 1
            else:
                #variable 
                pos_input = 0
                #click plus
                while True:
                    time.sleep(1)
                    try:
                        driver.find_element(By.XPATH,'/html/body/div[3]/div').click()
                    except:
                        try:
                            try:
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,f'/html/body/form/div[4]/div[3]/div/div[1]/div/div[1]/div/div[2]/table/tbody/tr[{pos_current_part+1}]/td[8]/button'))).click()
                            except:
                                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,f'/html/body/form/div[4]/div[3]/div/div[1]/div/div[1]/div/div[2]/table/tbody/tr[{pos_current_part+1}]/td[1]/button'))).click()
                        except:
                            break
                while True:
                    time.sleep(1)
                    try:
                        iframe = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="ifrDlgCommonCBM"]')))
                        break
                    except:
                        pass
                time.sleep(1)
                while True:
                    time.sleep(1)
                    try:
                        driver.switch_to.frame(iframe)
                        break
                    except:
                        pass
                time.sleep(1)
                #----------------------------------------------------------------------------
                # start time - end time
                if check_None(info[row][var_current_part][pos_input]) == False:
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[1]/td/span/table/tbody/tr[7]/td[2]/span[1]/input'))).clear()
                    time.sleep(0.2)
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[1]/td/span/table/tbody/tr[7]/td[2]/span[1]/input'))).send_keys(info[row][var_current_part][pos_input])
                    pos_input+=1
                    time.sleep(0.2)
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[1]/td/span/table/tbody/tr[7]/td[2]/span[2]/input'))).clear()
                    time.sleep(0.2)
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[1]/td/span/table/tbody/tr[7]/td[2]/span[2]/input'))).send_keys(info[row][var_current_part][pos_input])
                    pos_input+=1
                    time.sleep(0.1)
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[1]/td/span/table/tbody/tr[7]/td[1]/span[1]'))).click()
                    time.sleep(0.1)
                #----------------------------------------------------------------------------
                pos_combobox = 6
                error = 0
                #fill body form
                err = False
                res = True
                i = 1
                while(i<99 and res):
                    i+=1
                    try:
                        try:
                            try:
                                driver.find_element(By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/textarea').clear()
                                time.sleep(0.2)
                                driver.find_element(By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/textarea').send_keys(info[row][var_current_part][pos_input])  
                                pos_input+=1
                                error = 0
                                print(i,"textarea")
                            except:
                                driver.find_element(By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/input').clear()
                                time.sleep(0.2)
                                driver.find_element(By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/input').send_keys(info[row][var_current_part][pos_input])  
                                pos_input+=1
                                error = 0
                                print(i,"input")
                        except:
                            try:
                                #combobox
                                driver.find_element(By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/span/button').click()
                                time.sleep(1)
                                for k in range(1,5):    
                                    text = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH,f'/html/body/div[{pos_combobox}]/table/tbody/tr[{k}]/td/label'))).text
                                    if(text == info[row][var_current_part][pos_input]):
                                        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH,f'/html/body/div[{pos_combobox}]/table/tbody/tr[{k}]'))).click()
                                        pos_input+=1
                                        pos_combobox+=1
                                        error = 0
                                        print(i,"Combobox")
                                        break
                                    time.sleep(0.2)
                            except:
                                try:
                                    #fill value
                                    driver.find_element(By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/span/input[1]').clear()
                                    time.sleep(0.2)
                                    driver.find_element(By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/span/input[1]').send_keys(info[row][var_current_part][pos_input])
                                    pos_input+=1
                                    error = 0
                                    print(i,"fill value")
                                except:
                                    #begin checkbox
                                    list_checkbox = info[row][var_current_part][pos_input].split(',')
                                    count = 0
                                    checked = False
                                    for l in range(1,5):
                                        if(count == len(list_checkbox)): break
                                        try:
                                            for j in range(2,7,2):
                                                    text = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/table/tbody/tr[{l}]/td[{j}]/label'))).text
                                                    if(text == list_checkbox[count].strip()):
                                                        print("Dang click")
                                                        count +=1
                                                        WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH,f'/html/body/form/table/tbody/tr[2]/td[1]/div/div/table/tbody/tr[2]/td/span/table/tbody/tr[1]/td/div/div/table/tbody/tr[{i}]/td[2]/div/table/tbody/tr[{l}]/td[{j-1}]/div/div[2]/span'))).click()                                                        
                                                        checked =  True
                                                        print(i,"Checkbox")
                                                    time.sleep(0.1)
                                        except:
                                            if checked:
                                                print("Xong Checkbox")
                                                pos_input+=1
                                                error = 0
                                            temp = info[row][var_current_part][pos_input]
                                            print(temp)
                                            if len(temp.strip())==0 or temp == None or temp == 'None':
                                                print("End Line")
                                                res = False
                                            break
                                    #end checkbox
                                #----------------------------------------------------------------------------
                    except:
                        print(i,"None")
                        try:
                            temp = info[row][var_current_part][pos_input]
                            print(temp)
                            if len(temp.strip())==0 or temp == None or temp == "1":
                                print("End Line")
                                break
                        except:
                            break
                        error +=1
                        if error > 4:
                            # print("Lost.................")
                            # raise
                            err =  True
                            break
                if err:
                    sheet[row+3][array_column_part[-1]+var_current_part+1].value = "Chưa thực hiện"
                    ex.save(xpath_file)
                    continue                           
                else:
                    sheet[row+3][array_column_part[-1]+var_current_part+1].value = "Đã thực hiện"
                    ex.save(xpath_file)
                #luu
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/table/tbody/tr[1]/td/div/div[1]/button[2]'))).click()
                time.sleep(2)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/table/tbody/tr[1]/td/div/div[1]/button[4]'))).click()
                # take iframe
                while True:
                    time.sleep(1)
                    try:
                        iframe1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="ifrdlgChangeStatus"]')))
                        break
                    except:
                        pass

                time.sleep(2)
                while True:
                    time.sleep(1)
                    try:
                        driver.switch_to.frame(iframe1)
                        break
                    except:
                        pass
                time.sleep(1)
                while True:
                    time.sleep(1)
                    try:
                        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="formDlg:j_idt29"]/div[3]'))).click()
                        break
                    except:
                        pass
                time.sleep(0.5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/div[3]/div/ul/li[2]'))).click()
                time.sleep(0.5)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/center/button[2]'))).click()
                
                time.sleep(1)
                driver.get(link_cbm)
                time.sleep(1)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/div[4]/div[1]/div/table/tbody/tr/td[3]/input'))).send_keys(info[row][0])
                time.sleep(0.5)
                
                driver.find_element(By.XPATH,'/html/body/form/div[4]/div[1]/div/table/tbody/tr/td[4]/button/span[1]').click()
                time.sleep(1)
                while True:
                    time.sleep(1)
                    try:
                        driver.find_element(By.XPATH,'/html/body/div[3]/div').click()
                    except:
                        break
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,'/html/body/form/div[4]/div[2]/div/div/div[2]/table/tbody/tr/td[2]'))).click()
                time.sleep(2)
                # button lock                                                                            
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,f'/html/body/form/div[4]/div[3]/div/div[1]/div/div[1]/div/div[2]/table/tbody/tr[{pos_current_part+1}]/td[2]/button'))).click()
                while True:
                    time.sleep(1)
                    try:
                        driver.find_element(By.XPATH,'/html/body/div[3]/div').click()
                    except:
                        break
                try: 
                    driver.find_element(By.XPATH,f'/html/body/form/div[4]/div[3]/div/div[1]/div/div[1]/div/div[2]/table/tbody/tr[{pos_current_part+1}]/td[2]/button').click()
                    raise
                except: pass

                
def graphics():
    win = Tk()
    win.title("By Trương Tùng - XNCT Lai Châu")
    win.geometry('800x500')
    label = Label(win,text='Phần Mềm Tự Động Hóa',font=('arial',15))
    label.place(x=85,y=10)
    global str_sheet
    global checkbox_vars
    global checkbox_values
    checkbox_vars = []
    str_sheet = []
    checkbox_values = []
    
    def setup_initial_values():
        global str_sheet
        global checkbox_vars
        global checkbox_values
        

        # Xóa các checkbox cũ nếu có
        for checkbox in checkbox_values:
            checkbox.destroy()
        checkbox_values.clear()
        checkbox_vars.clear()
        # sheet | array_column_part | toltal_part
        # array = [[0,13,23,32,41,42,43],6]
        str_sheet = [
                    ['MC 22-35',[0,15,27,38,49,52,55]],
                     ['MC 110',[0,23,37,52,57,60,63]],
                     ['MBA',[0,4,54,84,89,97,102,108,120,145,156,159,163,169]],
                     ['MBA TD',[0,28,40,43,46]],
                     ['Tu',[0,4,11,22,31,35,38]],
                     ['Tu 22-35',[0,4,13,22,31,40,44,47]],
                     ['Tu 3 Pha',[0,9,22,37,40,43]],
                     ['TI',[0,4,15,24,28,31]],
                     ['TI 35-22',[0,4,19,28,32,35]],
                     ['DCL',[0,4,19,34,38,41]],
                     ['CSV',[0,5,13,27,35,38,41]],
                     ['CSV 2',[0,5,19,27,31,34]],
                     ['Cap Tong',[0,4,17,28,38,44,47]],
                     ['Ac Quy',[0,4,11,16,20,23]],
                     ['Day Dan',[0,4,8,26,29,32]],
                     ['Su 110',[0,15,18,21]],
                     ['MC 110 Hgis',[0,32,39,44,47,50]],
                     ['MC35kV ngoài trời',[0,14,29,34,37,40]],
                     ['Tụ bù 35',[0,15,27,38,49,52,55]],
                     ['CUỘN KHÁNG 22&35',[0,4,26,41,45,48]],
                     ['TỤ BÙ 22',[0,10,19,23,27,31],]
                     ]

        # Tạo lại các checkbox với danh sách str_sheet mới
        parameter = 0
        count = 0
        for i, sheet in enumerate(str_sheet): 
            var = IntVar(value=0)
            checkbox = Checkbutton(win, text=sheet[0], variable=var)
            checkbox.place(x=400 + parameter, y=50 + count * 30)  # Điều chỉnh vị trí của mỗi checkbox
            count+=1
            if i%12==0 and i != 0: 
                parameter+=200
                count = 0
            checkbox_values.append(checkbox)  # Thêm checkbox vào danh sách checkbox_values
            checkbox_vars.append(var)

        
    def event(object):
        global file_path
        if object == bt_ctn:
            if len(str_sheet) != len(checkbox_vars):
                messagebox.showinfo("Thông Báo", "Đã xảy ra lỗi vui lòng liên hệ coder!!!")
            try:
                user_text.get(),password_text.get()
                login(user_text.get(),password_text.get(),address_link_text.get(),file_path)
            except Exception as e:
                print(e)
                messagebox.showinfo("Thông Báo", "Đã xảy ra lỗi vui lòng kiểm tra!!!")
        elif object == bt_xpath:
            file_path = filedialog.askopenfilename(title="Chọn file")
            label1 = Label(win,text= file_path.split('/')[-1],font=('arial',10))
            label1.place(x=120,y=80)
        elif object == bt_update:
            try:
                chromedriver_autoinstaller.install()
                messagebox.showinfo("Thông Báo", "Đã cập nhật chromedrive thành công")
            except:
                messagebox.showinfo("Thông Báo", "Đã cập nhật chromedrive thất bại")
    
        
    label_path = Label(win,text='Địa chỉ link excel: ',font=('arial',10))
    label_path.place(x=10,y=50)
    bt_xpath = Button(win,text='Chọn File',font=('arial',10),padx=10,command=lambda: event(bt_xpath))
    bt_xpath.place(x=120,y=50)
    label_note = Label(win,text='(Y/C: Đúng dạng quy định sẵn)',font=('arial',10))
    label_note.place(x=210,y=50)


    label_path1 = Label(win,text='Chọn đường dẫn: ',font=('arial',10))
    label_path1.place(x=10,y=110)
    address_link_text = Combobox(win, font=('arial',10),width=28)
    address_link_text['value'] = ('https://pmis.npc.com.vn/PMIS_Web/tmsLogin.jsf','https://10.21.50.29/PMIS_Web/tmsLogin.jsf?faces-redirect=true')
    address_link_text.current(0)
    address_link_text.place(x=120,y=110)
    address_link_text.bind('<<ComboboxSelected>>',setup_initial_values())
    address_link_text.bind('<<ComboboxSelected>>', lambda event: setup_initial_values())

    label_path2 = Label(win,text='Tài khoản: ',font=('arial',10))
    label_path2.place(x=10,y=140)
    user_text = Entry(win, font=('arial',10),width=30,borderwidth=2)
    user_text.place(x=120,y=140)

    label_path3 = Label(win,text='Mật khẩu: ',font=('arial',10))
    label_path3.place(x=10,y=170)
    password_text = Entry(win, font=('arial',10),width=30,borderwidth=2)
    password_text.place(x=120,y=170)

    bt_ctn = Button(win,text='Tiếp Tục',font=('arial',10),padx=25,command=lambda: event(bt_ctn))
    bt_ctn.place(x=150,y=210)
    
    bt_update = Button(win,text='Cập Nhật Chrome',font=('arial',10),padx=35,command=lambda: event(bt_update))
    bt_update.place(x=600,y=10)
    win.mainloop()

def main():
    graphics()

if __name__ == '__main__':
    main()
